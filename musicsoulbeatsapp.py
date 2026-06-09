# # musicsoulbeatsapp.py
# # Complete Streamlit app (error-proof) with weekly Excel summary (Option B + choice 4)
# # Save this file as `musicsoulbeatsapp.py` and run with:
# #   streamlit run musicsoulbeatsapp.py
# #
# # Required installs (run once):
# # pip install streamlit opencv-python pandas numpy SpeechRecognition textblob emoji openpyxl matplotlib pillow
# # And (optional) for live mic: pip install pipwin -> pipwin install pyaudio (Windows)

# import streamlit as st
# import cv2
# import numpy as np
# import pandas as pd
# import os
# import pickle
# import webbrowser
# from datetime import datetime, timedelta
# from pathlib import Path
# import random
# import io
# import matplotlib.pyplot as plt
# from PIL import Image

# # optional imports
# try:
#     import speech_recognition as sr
#     SR_AVAILABLE = True
# except Exception:
#     SR_AVAILABLE = False

# try:
#     from textblob import TextBlob
#     TEXTBLOB_AVAILABLE = True
# except Exception:
#     TEXTBLOB_AVAILABLE = False

# # excel helper
# try:
#     import openpyxl
#     from openpyxl.drawing.image import Image as XLImage
#     OPENPYXL_AVAILABLE = True
# except Exception:
#     OPENPYXL_AVAILABLE = False

# # ---------------- Configuration ----------------
# st.set_page_config(page_title="SoulBeat — Full App + Weekly Excel", layout="wide")

# DATA_DIR = Path("soul_data")
# DATA_DIR.mkdir(exist_ok=True)
# DESC_DB = DATA_DIR / "face_descs.pkl"
# DIARY_CSV = DATA_DIR / "mood_diary.csv"
# WEEKLY_XLSX = DATA_DIR / "weekly_summary.xlsx"

# DIARY_COLUMNS = ["timestamp", "username", "mood", "source", "youtube_url", "wellness", "login_count"]

# # ORB + matcher
# ORB = cv2.ORB_create(500)
# BF = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)

# # Playlists - put distinct URLs per mood
# YOUTUBE_PLAYLISTS = {
#     "happy": [
#         "https://www.youtube.com/watch?v=3JWTaaS7LdU",
#         "https://www.youtube.com/watch?v=d-diB65scQU",
#         "https://www.youtube.com/watch?v=2Vv-BfVoq4g"
#     ],
#     "neutral": [
#         "https://www.youtube.com/watch?v=fLexgOxsZu0",
#         "https://www.youtube.com/watch?v=5qap5aO4i9A",
#         "https://www.youtube.com/watch?v=DWcJFNfaw9c"
#     ],
#     "sad": [
#         "https://www.youtube.com/watch?v=hTWKbfoikeg",
#         "https://www.youtube.com/watch?v=RB-RcX5DS5A",
#         "https://www.youtube.com/watch?v=2XU0oxnq2qU"
#     ]
# }

# WELLNESS_SUGGESTIONS = {
#     "happy": [
#         "Keep your energy — try a focused 15-min productive sprint.",
#         "Share your positivity: message a friend or write a gratitude note.",
#         "Short stretching + deep breaths to sustain mood."
#     ],
#     "neutral": [
#         "Take a 10-minute walk to refresh your mind.",
#         "Do a 5-minute breathing exercise to improve focus.",
#         "Hydrate and have a healthy snack."
#     ],
#     "sad": [
#         "Try 10 minutes of guided breathing or meditation.",
#         "Do gentle stretching/yoga and drink warm water.",
#         "Write one small positive thing and focus on it for 5 minutes."
#     ]
# }

# # ---------- Utility functions ----------

# def ensure_diary_schema():
#     if not DIARY_CSV.exists():
#         df = pd.DataFrame(columns=DIARY_COLUMNS)
#         df.to_csv(DIARY_CSV, index=False)
#         return df
#     try:
#         df = pd.read_csv(DIARY_CSV)
#         missing = [c for c in DIARY_COLUMNS if c not in df.columns]
#         if missing:
#             for c in missing:
#                 df[c] = 0 if c == "login_count" else ""
#             df = df.reindex(columns=DIARY_COLUMNS)
#             df.to_csv(DIARY_CSV, index=False)
#         return df
#     except Exception:
#         backup = DIARY_CSV.with_suffix('.bak.csv')
#         try:
#             DIARY_CSV.replace(backup)
#         except Exception:
#             pass
#         df = pd.DataFrame(columns=DIARY_COLUMNS)
#         df.to_csv(DIARY_CSV, index=False)
#         return df


# def load_desc_db():
#     if DESC_DB.exists():
#         try:
#             with open(DESC_DB, 'rb') as f:
#                 return pickle.load(f)
#         except Exception:
#             return {}
#     return {}


# def save_desc_db(db):
#     with open(DESC_DB, 'wb') as f:
#         pickle.dump(db, f)


# def compute_orb_descriptor(gray_face):
#     try:
#         kps, des = ORB.detectAndCompute(gray_face, None)
#         return des
#     except Exception:
#         return None


# def match_descriptor(des, db, min_matches=7):
#     if des is None:
#         return None, 0
#     best_user, best_score = None, 0
#     for user, samples in db.items():
#         for stored in samples:
#             if stored is None:
#                 continue
#             try:
#                 matches = BF.match(des, stored)
#             except Exception:
#                 continue
#             good = [m for m in matches if m.distance < 60]
#             if len(good) > best_score:
#                 best_score = len(good)
#                 best_user = user
#     if best_score >= min_matches:
#         return best_user, best_score
#     return None, best_score


# def detect_faces_haar(frame):
#     gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
#     face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
#     faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(80,80))
#     boxes = []
#     for (x,y,w,h) in faces:
#         boxes.append((x,y,x+w,y+h))
#     return boxes


# def estimate_mood_from_face(face_bgr):
#     if face_bgr is None or face_bgr.size == 0:
#         return 'neutral', 0.5
#     gray = cv2.cvtColor(face_bgr, cv2.COLOR_BGR2GRAY)
#     hsv = cv2.cvtColor(face_bgr, cv2.COLOR_BGR2HSV)
#     mean_v = float(np.mean(hsv[:,:,2]))
#     smile_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_smile.xml')
#     smiles = smile_cascade.detectMultiScale(gray, scaleFactor=1.7, minNeighbors=22)
#     if len(smiles) > 0 or mean_v > 150:
#         return 'happy', round(mean_v/255.0,2)
#     if mean_v < 70:
#         return 'sad', round(1 - mean_v/255.0,2)
#     return 'neutral', round(abs(mean_v/255.0 - 0.5),2)


# def text_to_mood(text):
#     if not text or not TEXTBLOB_AVAILABLE:
#         t = (text or '').lower()
#         if any(w in t for w in ['sad','unhappy','depress','cry','lonely']):
#             return 'sad'
#         if any(w in t for w in ['happy','joy','glad','love','smile']):
#             return 'happy'
#         return 'neutral'
#     try:
#         pol = TextBlob(text).sentiment.polarity
#         if pol > 0.25:
#             return 'happy'
#         if pol < -0.25:
#             return 'sad'
#         return 'neutral'
#     except Exception:
#         return 'neutral'


# def choose_unique_youtube(username, mood):
#     df = ensure_diary_schema()
#     if df.empty:
#         played = set()
#     else:
#         try:
#             played = set(df[df['username']==username]['youtube_url'].astype(str).tolist())
#         except Exception:
#             played = set()
#     candidates = YOUTUBE_PLAYLISTS.get(mood, [])
#     remaining = [u for u in candidates if u not in played]
#     if remaining:
#         return random.choice(remaining)
#     return random.choice(candidates) if candidates else None


# def pick_wellness(username, mood):
#     df = ensure_diary_schema()
#     count = 0
#     if not df.empty:
#         try:
#             count = int(df[df['username']==username].shape[0])
#         except Exception:
#             count = 0
#     pool = WELLNESS_SUGGESTIONS.get(mood, ['Take a short break and breathe.'])
#     return pool[count % len(pool)]


# def log_entry(username, mood, source, youtube_url, wellness):
#     df = ensure_diary_schema()
#     try:
#         login_count = int(df[df['username']==username].shape[0]) + 1 if not df.empty else 1
#     except Exception:
#         login_count = 1
#     row = {
#         'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
#         'username': username,
#         'mood': mood,
#         'source': source,
#         'youtube_url': youtube_url,
#         'wellness': wellness,
#         'login_count': login_count
#     }
#     df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
#     df.to_csv(DIARY_CSV, index=False)

# # ---------- Weekly Excel generation ----------

# def generate_weekly_excel(output_path=WEEKLY_XLSX, days=7):
#     df = ensure_diary_schema()
#     if df.empty:
#         # create minimal workbook
#         with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
#             pd.DataFrame(columns=DIARY_COLUMNS).to_excel(writer, sheet_name='RawLogs', index=False)
#         return output_path

#     cutoff = pd.Timestamp.now() - pd.Timedelta(days=days)
#     df['timestamp_dt'] = pd.to_datetime(df['timestamp'], errors='coerce')
#     recent = df[df['timestamp_dt'] >= cutoff].copy()
#     if recent.empty:
#         recent = pd.DataFrame(columns=df.columns)

#     # Sheet1: Raw Logs
#     # Sheet2: Mood Summary (counts per day)
#     # Sheet3: Charts (we will insert PNG images)
#     # Sheet4: Recommendations Summary
#     # Prepare mood summary
#     if not recent.empty:
#         recent['date'] = recent['timestamp_dt'].dt.date
#         mood_counts = recent.groupby(['date','mood']).size().unstack(fill_value=0)
#         mood_daily = mood_counts.reset_index()
#         mood_totals = recent['mood'].value_counts().rename_axis('mood').reset_index(name='count')
#     else:
#         mood_daily = pd.DataFrame()
#         mood_totals = pd.DataFrame(columns=['mood','count'])

#     # Recommendations summary
#     if not recent.empty:
#         recs = recent.groupby('wellness').size().reset_index(name='count').sort_values('count', ascending=False)
#     else:
#         recs = pd.DataFrame(columns=['wellness','count'])

#     # Write sheets
#     if not OPENPYXL_AVAILABLE:
#         # use pandas ExcelWriter but require openpyxl; if not available, fall back to CSV exports
#         panic = output_path.with_suffix('.csv')
#         recent.to_csv(panic, index=False)
#         return panic

#     with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
#         recent.drop(columns=['timestamp_dt'], errors='ignore').to_excel(writer, sheet_name='RawLogs', index=False)
#         mood_daily.to_excel(writer, sheet_name='MoodDaily', index=False)
#         mood_totals.to_excel(writer, sheet_name='MoodTotals', index=False)
#         recs.to_excel(writer, sheet_name='Recommendations', index=False)

#     # Generate charts as images and embed into workbook
#     try:
#         # Chart 1: Mood totals bar
#         fig1, ax1 = plt.subplots(figsize=(6,4))
#         if not mood_totals.empty:
#             ax1.bar(mood_totals['mood'], mood_totals['count'])
#             ax1.set_title('Mood distribution (last {} days)'.format(days))
#             ax1.set_ylabel('Count')
#         else:
#             ax1.text(0.5,0.5,'No data',ha='center')
#         img1 = output_path.parent / 'chart1.png'
#         fig1.savefig(img1)
#         plt.close(fig1)

#         # Chart 2: Mood trend per day (stacked)
#         fig2, ax2 = plt.subplots(figsize=(8,4))
#         if not mood_daily.empty:
#             mood_daily.set_index('date').plot(kind='bar', stacked=True, ax=ax2)
#             ax2.set_title('Mood counts by day')
#         else:
#             ax2.text(0.5,0.5,'No data',ha='center')
#         img2 = output_path.parent / 'chart2.png'
#         fig2.savefig(img2)
#         plt.close(fig2)

#         # Insert images into workbook
#         wb = openpyxl.load_workbook(output_path)
#         ws = wb.create_sheet('Charts')
#         img_obj1 = XLImage(str(img1))
#         img_obj2 = XLImage(str(img2))
#         ws.add_image(img_obj1, 'A1')
#         ws.add_image(img_obj2, 'A20')
#         wb.save(output_path)

#         # cleanup images
#         try:
#             img1.unlink()
#             img2.unlink()
#         except Exception:
#             pass
#     except Exception as e:
#         # if chart generation fails, still return workbook
#         print('Chart embed error:', e)

#     return output_path

# # ---------- Streamlit UI ----------
# st.title('SoulBeat — Full app with Weekly Excel Export')

# st.sidebar.header('User & Settings')
# username = st.sidebar.text_input('Username (for diary):', value=st.session_state.get('username',''))
# if username:
#     st.session_state['username']=username

# mic_enabled = st.sidebar.checkbox('Enable microphone (requires PyAudio)', value=False)

# st.sidebar.markdown('Run with: streamlit run musicsoulbeatsapp.py')

# col1, col2 = st.columns([1,2])

# with col1:
#     st.header('Manual Controls')
#     emoji_map = {'😞':'sad','😐':'neutral','🙂':'happy','😊':'happy','😡':'sad'}
#     emoji_choice = st.select_slider('Pick an emoji', options=list(emoji_map.keys()), value='😐')
#     manual_mood = emoji_map.get(emoji_choice,'neutral')
#     typed = st.text_input('Or type mood (happy/neutral/sad):','')
#     override_mood = typed.strip().lower() if typed.strip().lower() in {'happy','neutral','sad'} else None

#     st.markdown('---')
#     st.header('Speech Input')
#     if not SR_AVAILABLE:
#         st.warning('SpeechRecognition package not installed — install with `pip install SpeechRecognition`. Audio upload still works.')

#     if mic_enabled and SR_AVAILABLE:
#         if st.button('Record 5s from mic'):
#             try:
#                 r = sr.Recognizer()
#                 with sr.Microphone() as source:
#                     st.info('Recording...')
#                     r.adjust_for_ambient_noise(source, duration=0.4)
#                     audio = r.record(source, duration=5)
#                 st.success('Recording done — transcribing...')
#                 text = r.recognize_google(audio)
#                 st.write('Transcribed:', text)
#                 mood = text_to_mood(text)
#                 st.write('Mood:', mood)
#                 yt = choose_unique_youtube(username or 'anonymous', mood)
#                 wellness = pick_wellness(username or 'anonymous', mood)
#                 if yt:
#                     log_entry(username or 'anonymous', mood, 'mic_speech', yt, wellness)
#                     st.video(yt)
#                     st.success('Playing')
#                     st.info('Wellness: ' + wellness)
#                 else:
#                     st.warning('No yt configured')
#             except Exception as e:
#                 st.error('Mic capture error (PyAudio?) ' + str(e))
#     else:
#         if mic_enabled and not SR_AVAILABLE:
#             st.info('Mic requested but SpeechRecognition not available')

#     uploaded = st.file_uploader('Upload audio (wav preferred) for speech-to-mood', type=['wav','mp3','m4a'])
#     if uploaded and st.button('Transcribe upload and play'):
#         if not SR_AVAILABLE:
#             st.error('SpeechRecognition not installed')
#         else:
#             tmp = DATA_DIR / f"tmp_{datetime.now().strftime('%Y%m%d%H%M%S')}"

#             ext = uploaded.name.split('.')[-1]
#             tmp = tmp.with_suffix('.' + ext)
#             with open(tmp, 'wb') as f:
#                 f.write(uploaded.read())
#             r = sr.Recognizer()
#             try:
#                 with sr.AudioFile(str(tmp)) as source:
#                     audio = r.record(source)
#                 text = r.recognize_google(audio)
#                 st.write('Transcribed:', text)
#                 mood = text_to_mood(text)
#                 st.write('Mood:', mood)
#                 yt = choose_unique_youtube(username or 'anonymous', mood)
#                 wellness = pick_wellness(username or 'anonymous', mood)
#                 if yt:
#                     log_entry(username or 'anonymous', mood, 'upload_audio', yt, wellness)
#                     st.video(yt)
#                     st.success('Playing')
#                     st.info('Wellness: ' + wellness)
#                 else:
#                     st.warning('No yt configured for this mood')
#             except sr.UnknownValueError:
#                 st.error('Could not understand audio')
#             except Exception as e:
#                 st.error('Transcription error: ' + str(e))
#             finally:
#                 try:
#                     tmp.unlink()
#                 except Exception:
#                     pass

#     st.markdown('---')
#     st.header('Manual Playback')
#     if st.button('Play for manual mood'):
#         final_mood = override_mood or manual_mood
#         yt = choose_unique_youtube(username or 'anonymous', final_mood)
#         wellness = pick_wellness(username or 'anonymous', final_mood)
#         if yt:
#             log_entry(username or 'anonymous', final_mood, 'manual', yt, wellness)
#             st.video(yt)
#             st.success('Playing')
#             st.info('Wellness: ' + wellness)
#         else:
#             st.warning('No yt configured')

# with col2:
#     st.header('Face capture & recognition (login)')
#     st.write('Capture photo and the app will detect face, recognize user, detect mood, play unique song and log entry')
#     img = st.camera_input('Capture photo')
#     if img is not None:
#         arr = np.frombuffer(img.getvalue(), np.uint8)
#         frame = cv2.imdecode(arr, cv2.IMREAD_COLOR)
#         st.image(frame, channels='BGR')
#         boxes = detect_faces_haar(frame)
#         if not boxes:
#             st.warning('No face detected')
#         else:
#             x1,y1,x2,y2 = boxes[0]
#             face = frame[y1:y2, x1:x2]
#             st.image(face, channels='BGR', caption='Face crop')
#             gray = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
#             des = compute_orb_descriptor(gray)
#             db = load_desc_db()
#             matched, score = match_descriptor(des, db, min_matches=6)
#             if matched is None:
#                 reg_name = username.strip() if username and username.strip() else f"anon_{datetime.now().strftime('%Y%m%d%H%M%S')}"
#                 db.setdefault(reg_name, []).append(des)
#                 save_desc_db(db)
#                 st.info(f'Registered descriptor for {reg_name}')
#                 userkey = reg_name
#             else:
#                 userkey = matched
#                 db.setdefault(userkey, []).append(des)
#                 save_desc_db(db)
#                 st.success(f'Recognized {userkey} (score {score})')
#             mood, conf = estimate_mood_from_face(face)
#             st.write(f'Detected mood: {mood} (conf {conf})')
#             yt = choose_unique_youtube(userkey, mood)
#             wellness = pick_wellness(userkey, mood)
#             if yt:
#                 log_entry(userkey, mood, 'face_capture', yt, wellness)
#                 st.video(yt)
#                 st.success('Playing unique song')
#                 st.info('Wellness: ' + wellness)
#             else:
#                 st.warning('No yt configured for this mood')

#     st.markdown('---')
#     st.header('Diary & Playlists')
#     df = ensure_diary_schema()
#     if df.empty:
#         st.info('Diary empty')
#     else:
#         st.dataframe(df.sort_values(by='timestamp', ascending=False).reset_index(drop=True))

#     st.markdown('Playlist status for your user')
#     user_key = username if username else 'anonymous'
#     try:
#         played_urls = set(df[df['username']==user_key]['youtube_url'].astype(str).tolist()) if not df.empty else set()
#     except Exception:
#         played_urls = set()
#     st.write(f'Previously played for {user_key}: {len(played_urls)}')
#     for mood, urls in YOUTUBE_PLAYLISTS.items():
#         st.write(f'- {mood} ({len(urls)})')
#         for u in urls:
#             tag = ' (played)' if u in played_urls else ''
#             st.markdown(f'   - {u}{tag}')

# st.markdown('---')
# st.header('Weekly Excel Summary / Export')
# if st.button('Generate weekly Excel (7 days)'):
#     output = generate_weekly_excel()
#     st.success(f'Weekly workbook created: {output}')
#     if output.exists():
#         with open(output, 'rb') as f:
#             st.download_button('Download weekly_summary.xlsx', f, file_name=output.name)

# st.markdown('Notes: For mic recording install PyAudio (Windows: pip install pipwin; pipwin install pyaudio).')

# musicsoulbeatsapp_fixed.py
# Fixed, error-proof Streamlit app with weekly Excel summary
# Save as musicsoulbeatsapp.py and run:
#    streamlit run musicsoulbeatsapp.py

import streamlit as st
import cv2
import numpy as np
import pandas as pd
import os
import pickle
from datetime import datetime, timedelta
from pathlib import Path
import random
import io
import matplotlib.pyplot as plt
from PIL import Image

# optional imports
try:
    import speech_recognition as sr
    SR_AVAILABLE = True
except Exception:
    SR_AVAILABLE = False

# check for PyAudio specifically (required for live mic)
try:
    import pyaudio  # noqa: F401
    PYAUDIO_AVAILABLE = True
except Exception:
    PYAUDIO_AVAILABLE = False

try:
    from textblob import TextBlob
    TEXTBLOB_AVAILABLE = True
except Exception:
    TEXTBLOB_AVAILABLE = False

# excel helper
try:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

st.set_page_config(page_title="SoulBeat — Full App + Weekly Excel", layout="wide")

# Data paths (keep out of OneDrive if possible)
DATA_DIR = Path("soul_data")
DATA_DIR.mkdir(exist_ok=True)
DESC_DB = DATA_DIR / "face_descs.pkl"
DIARY_CSV = DATA_DIR / "mood_diary.csv"
WEEKLY_XLSX = DATA_DIR / "weekly_summary.xlsx"

DIARY_COLUMNS = ["timestamp", "username", "mood", "source", "youtube_url", "wellness", "login_count"]

# ORB + matcher
ORB = cv2.ORB_create(500)
BF = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)

# Playlists - distinct URLs per mood
YOUTUBE_PLAYLISTS = {
    "happy": [
        "https://www.youtube.com/watch?v=3JWTaaS7LdU",
        "https://www.youtube.com/watch?v=d-diB65scQU",
        "https://www.youtube.com/watch?v=2Vv-BfVoq4g"
    ],
    "neutral": [
        "https://www.youtube.com/watch?v=fLexgOxsZu0",
        "https://www.youtube.com/watch?v=5qap5aO4i9A",
        "https://www.youtube.com/watch?v=DWcJFNfaw9c"
    ],
    "sad": [
        "https://www.youtube.com/watch?v=hTWKbfoikeg",
        "https://www.youtube.com/watch?v=RB-RcX5DS5A",
        "https://www.youtube.com/watch?v=2XU0oxnq2qU"
    ]
}

WELLNESS_SUGGESTIONS = {
    "happy": [
        "Keep your energy — try a focused 15-min productive sprint.",
        "Share your positivity: message a friend or write a gratitude note.",
        "Short stretching + deep breaths to sustain mood."
    ],
    "neutral": [
        "Take a 10-minute walk to refresh your mind.",
        "Do a 5-minute breathing exercise to improve focus.",
        "Hydrate and have a healthy snack."
    ],
    "sad": [
        "Try 10 minutes of guided breathing or meditation.",
        "Do gentle stretching/yoga and drink warm water.",
        "Write one small positive thing and focus on it for 5 minutes."
    ]
}

# ---------- Utility functions ----------

def ensure_diary_schema():
    if not DIARY_CSV.exists():
        df = pd.DataFrame(columns=DIARY_COLUMNS)
        df.to_csv(DIARY_CSV, index=False)
        return df
    try:
        df = pd.read_csv(DIARY_CSV)
        missing = [c for c in DIARY_COLUMNS if c not in df.columns]
        if missing:
            for c in missing:
                df[c] = 0 if c == "login_count" else ""
            df = df.reindex(columns=DIARY_COLUMNS)
            df.to_csv(DIARY_CSV, index=False)
        return df
    except Exception:
        # recreate clean file
        df = pd.DataFrame(columns=DIARY_COLUMNS)
        df.to_csv(DIARY_CSV, index=False)
        return df


def load_desc_db():
    if DESC_DB.exists():
        try:
            with open(DESC_DB, 'rb') as f:
                return pickle.load(f)
        except Exception:
            return {}
    return {}


def save_desc_db(db):
    with open(DESC_DB, 'wb') as f:
        pickle.dump(db, f)


def compute_orb_descriptor(gray_face):
    if gray_face is None or gray_face.size == 0:
        return None
    try:
        # resize small faces to stable size for better descriptors
        h,w = gray_face.shape[:2]
        if h < 80 or w < 80:
            gray_face = cv2.resize(gray_face, (100,100))
        kps, des = ORB.detectAndCompute(gray_face, None)
        return des
    except Exception:
        return None


def match_descriptor(des, db, min_matches=6):
    if des is None:
        return None, 0
    best_user, best_score = None, 0
    for user, samples in db.items():
        for stored in samples:
            if stored is None:
                continue
            try:
                matches = BF.match(des, stored)
            except Exception:
                continue
            good = [m for m in matches if m.distance < 60]
            if len(good) > best_score:
                best_score = len(good)
                best_user = user
    if best_score >= min_matches:
        return best_user, best_score
    return None, best_score


def detect_faces_haar(frame):
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=4, minSize=(60,60))
    boxes = []
    for (x,y,w,h) in faces:
        boxes.append((x,y,x+w,y+h))
    return boxes


def estimate_mood_from_face(face_bgr):
    # more robust: use brightness + simple smile detection with relaxed params
    try:
        if face_bgr is None or face_bgr.size == 0:
            return 'neutral', 0.5
        face = face_bgr.copy()
        h,w = face.shape[:2]
        if h < 80 or w < 80:
            face = cv2.resize(face, (160,160))
        gray = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
        hsv = cv2.cvtColor(face, cv2.COLOR_BGR2HSV)
        mean_v = float(np.mean(hsv[:,:,2]))
        smile_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_smile.xml')
        # relaxed parameters to detect smiles more reliably
        smiles = smile_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=10, minSize=(20,20))
        # brightness thresholds tuned
        if len(smiles) > 0 or mean_v > 140:
            conf = min(1.0, mean_v/255.0 + 0.2)
            return 'happy', round(conf, 2)
        if mean_v < 90:
            conf = min(1.0, (90-mean_v)/90)
            return 'sad', round(conf, 2)
        return 'neutral', round(abs(mean_v/255.0 - 0.5), 2)
    except Exception:
        return 'neutral', 0.5


def text_to_mood(text):
    if not text:
        return 'neutral'
    if TEXTBLOB_AVAILABLE:
        try:
            pol = TextBlob(text).sentiment.polarity
            if pol > 0.2:
                return 'happy'
            if pol < -0.2:
                return 'sad'
            return 'neutral'
        except Exception:
            pass
    t = text.lower()
    if any(w in t for w in ['sad','unhappy','depress','cry','lonely','angry','mad']):
        return 'sad'
    if any(w in t for w in ['happy','joy','glad','love','smile']):
        return 'happy'
    return 'neutral'


def choose_unique_youtube(username, mood):
    df = ensure_diary_schema()
    if df.empty:
        played = set()
    else:
        try:
            played = set(df[df['username']==username]['youtube_url'].astype(str).tolist())
        except Exception:
            played = set()
    candidates = YOUTUBE_PLAYLISTS.get(mood, [])
    remaining = [u for u in candidates if u not in played]
    if remaining:
        return random.choice(remaining)
    return random.choice(candidates) if candidates else None


def pick_wellness(username, mood):
    df = ensure_diary_schema()
    count = 0
    if not df.empty:
        try:
            count = int(df[df['username']==username].shape[0])
        except Exception:
            count = 0
    pool = WELLNESS_SUGGESTIONS.get(mood, ['Take a short break and breathe.'])
    return pool[count % len(pool)]


def log_entry(username, mood, source, youtube_url, wellness):
    df = ensure_diary_schema()
    try:
        login_count = int(df[df['username']==username].shape[0]) + 1 if not df.empty else 1
    except Exception:
        login_count = 1
    row = {
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'username': username,
        'mood': mood,
        'source': source,
        'youtube_url': youtube_url,
        'wellness': wellness,
        'login_count': login_count
    }
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    df.to_csv(DIARY_CSV, index=False)

# ---------- Weekly Excel generation ----------

def generate_weekly_excel(output_path=WEEKLY_XLSX, days=7):
    df = ensure_diary_schema()
    if df.empty:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            pd.DataFrame(columns=DIARY_COLUMNS).to_excel(writer, sheet_name='RawLogs', index=False)
        return output_path

    cutoff = pd.Timestamp.now() - pd.Timedelta(days=days)
    df['timestamp_dt'] = pd.to_datetime(df['timestamp'], errors='coerce')
    recent = df[df['timestamp_dt'] >= cutoff].copy()
    if recent.empty:
        recent = pd.DataFrame(columns=df.columns)

    if not recent.empty:
        recent['date'] = recent['timestamp_dt'].dt.date
        mood_counts = recent.groupby(['date','mood']).size().unstack(fill_value=0)
        mood_daily = mood_counts.reset_index()
        mood_totals = recent['mood'].value_counts().rename_axis('mood').reset_index(name='count')
    else:
        mood_daily = pd.DataFrame()
        mood_totals = pd.DataFrame(columns=['mood','count'])

    if not recent.empty:
        recs = recent.groupby('wellness').size().reset_index(name='count').sort_values('count', ascending=False)
    else:
        recs = pd.DataFrame(columns=['wellness','count'])

    if not OPENPYXL_AVAILABLE:
        fallback = output_path.with_suffix('.csv')
        recent.to_csv(fallback, index=False)
        return fallback

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        recent.drop(columns=['timestamp_dt'], errors='ignore').to_excel(writer, sheet_name='RawLogs', index=False)
        mood_daily.to_excel(writer, sheet_name='MoodDaily', index=False)
        mood_totals.to_excel(writer, sheet_name='MoodTotals', index=False)
        recs.to_excel(writer, sheet_name='Recommendations', index=False)

    try:
        # Chart 1
        fig1, ax1 = plt.subplots(figsize=(6,4))
        if not mood_totals.empty:
            ax1.bar(mood_totals['mood'], mood_totals['count'])
            ax1.set_title('Mood distribution (last {} days)'.format(days))
            ax1.set_ylabel('Count')
        else:
            ax1.text(0.5,0.5,'No data',ha='center')
        img1 = output_path.parent / 'chart1.png'
        fig1.savefig(img1)
        plt.close(fig1)

        # Chart 2
        fig2, ax2 = plt.subplots(figsize=(8,4))
        if not mood_daily.empty:
            mood_daily.set_index('date').plot(kind='bar', stacked=True, ax=ax2)
            ax2.set_title('Mood counts by day')
        else:
            ax2.text(0.5,0.5,'No data',ha='center')
        img2 = output_path.parent / 'chart2.png'
        fig2.savefig(img2)
        plt.close(fig2)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.create_sheet('Charts')
        img_obj1 = XLImage(str(img1))
        img_obj2 = XLImage(str(img2))
        ws.add_image(img_obj1, 'A1')
        ws.add_image(img_obj2, 'A25')
        wb.save(output_path)

        try:
            img1.unlink()
            img2.unlink()
        except Exception:
            pass
    except Exception:
        pass

    return output_path

# ---------- Streamlit UI ----------
st.title('SoulBeat — Full app with Weekly Excel Export (Fixed)')

st.sidebar.header('User & Settings')
# use session_state key directly so username persists reliably
username = st.sidebar.text_input('Username (for diary):', key='username')
if not username:
    username = 'anonymous'

mic_enabled = st.sidebar.checkbox('Enable microphone (requires PyAudio)', value=False, key='mic_enabled')

st.sidebar.markdown('Run with: streamlit run musicsoulbeatsapp.py')

col1, col2 = st.columns([1,2])

with col1:
    st.header('Manual Controls')
    emoji_map = {'😞':'sad','😐':'neutral','🙂':'happy','😊':'happy','😡':'sad'}
    emoji_choice = st.select_slider('Pick an emoji', options=list(emoji_map.keys()), value='😐', key='emoji_slider')
    manual_mood = emoji_map.get(emoji_choice,'neutral')
    typed = st.text_input('Or type mood (happy/neutral/sad):','', key='typed_mood')
    override_mood = typed.strip().lower() if typed.strip().lower() in {'happy','neutral','sad'} else None

    st.markdown('---')
    st.header('Speech Input')
    if not SR_AVAILABLE:
        st.warning('SpeechRecognition package not installed — install with `pip install SpeechRecognition`. Audio upload still works.')
    elif mic_enabled and not PYAUDIO_AVAILABLE:
        st.info('PyAudio not available — install for live mic: pip install pipwin ; pipwin install pyaudio (Windows)')

    # Mic recording button (only run if SR_AVAILABLE and PyAudio available)
    if st.button('Record 5s from mic', key='record_mic'):
        if not SR_AVAILABLE or not PYAUDIO_AVAILABLE:
            st.error('Microphone recording requires SpeechRecognition + PyAudio.')
        else:
            try:
                r = sr.Recognizer()
                with sr.Microphone() as source:
                    st.info('Recording...')
                    r.adjust_for_ambient_noise(source, duration=0.4)
                    audio = r.record(source, duration=5)
                st.success('Recording done — transcribing...')
                text = r.recognize_google(audio)
                st.write('Transcribed:', text)
                mood = text_to_mood(text)
                st.write('Mood:', mood)
                yt = choose_unique_youtube(username or 'anonymous', mood)
                wellness = pick_wellness(username or 'anonymous', mood)
                if yt:
                    log_entry(username or 'anonymous', mood, 'mic_speech', yt, wellness)
                    st.video(yt)
                    st.success('Playing')
                    st.info('Wellness: ' + wellness)
                else:
                    st.warning('No yt configured')
            except Exception as e:
                st.error('Mic capture/transcription error: ' + str(e))

    # Uploaded audio: allow processing when button pressed
    uploaded = st.file_uploader('Upload audio (wav preferred) for speech-to-mood', type=['wav','mp3','m4a'], key='upload_audio')
    if uploaded is not None and st.button('Transcribe upload and play', key='transcribe_upload'):
        if not SR_AVAILABLE:
            st.error('SpeechRecognition not installed')
        else:
            tmp = DATA_DIR / f"tmp_{datetime.now().strftime('%Y%m%d%H%M%S')}"
            ext = uploaded.name.split('.')[-1]
            tmp = tmp.with_suffix('.' + ext)
            with open(tmp, 'wb') as f:
                f.write(uploaded.read())
            r = sr.Recognizer()
            try:
                with sr.AudioFile(str(tmp)) as source:
                    audio = r.record(source)
                text = r.recognize_google(audio)
                st.write('Transcribed:', text)
                mood = text_to_mood(text)
                st.write('Mood:', mood)
                yt = choose_unique_youtube(username or 'anonymous', mood)
                wellness = pick_wellness(username or 'anonymous', mood)
                if yt:
                    log_entry(username or 'anonymous', mood, 'upload_audio', yt, wellness)
                    st.video(yt)
                    st.success('Playing')
                    st.info('Wellness: ' + wellness)
                else:
                    st.warning('No yt configured for this mood')
            except sr.UnknownValueError:
                st.error('Could not understand audio')
            except Exception as e:
                st.error('Transcription error: ' + str(e))
            finally:
                try:
                    tmp.unlink()
                except Exception:
                    pass

    st.markdown('---')
    st.header('Manual Playback')
    if st.button('Play for manual mood', key='manual_play'):
        final_mood = override_mood or manual_mood
        yt = choose_unique_youtube(username or 'anonymous', final_mood)
        wellness = pick_wellness(username or 'anonymous', final_mood)
        if yt:
            log_entry(username or 'anonymous', final_mood, 'manual', yt, wellness)
            st.video(yt)
            st.success('Playing')
            st.info('Wellness: ' + wellness)
        else:
            st.warning('No yt configured')

with col2:
    st.header('Face capture & recognition (login)')
    st.write('Capture photo and the app will detect face, recognize user, detect mood, play unique song and log entry')
    img = st.camera_input('Capture photo', key='camera')
    if img is not None:
        arr = np.frombuffer(img.getvalue(), np.uint8)
        frame = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        st.image(frame, channels='BGR')
        boxes = detect_faces_haar(frame)
        if not boxes:
            st.warning('No face detected')
        else:
            x1,y1,x2,y2 = boxes[0]
            # clamp coords
            x1, y1 = max(0,x1), max(0,y1)
            x2, y2 = min(frame.shape[1], x2), min(frame.shape[0], y2)
            face = frame[y1:y2, x1:x2]
            st.image(face, channels='BGR', caption='Face crop')
            gray = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
            des = compute_orb_descriptor(gray)
            db = load_desc_db()
            matched, score = match_descriptor(des, db, min_matches=6)
            if matched is None:
                reg_name = username.strip() if username and username.strip() else f"anon_{datetime.now().strftime('%Y%m%d%H%M%S')}"
                db.setdefault(reg_name, []).append(des)
                save_desc_db(db)
                st.info(f'Registered descriptor for {reg_name}')
                userkey = reg_name
            else:
                userkey = matched
                db.setdefault(userkey, []).append(des)
                save_desc_db(db)
                st.success(f'Recognized {userkey} (score {score})')
            mood, conf = estimate_mood_from_face(face)
            st.write(f'Detected mood: {mood} (conf {conf})')
            yt = choose_unique_youtube(userkey, mood)
            wellness = pick_wellness(userkey, mood)
            if yt:
                log_entry(userkey, mood, 'face_capture', yt, wellness)
                st.video(yt)
                st.success('Playing unique song')
                st.info('Wellness: ' + wellness)
            else:
                st.warning('No yt configured for this mood')

    st.markdown('---')
    st.header('Diary & Playlists')
    df = ensure_diary_schema()
    if df.empty:
        st.info('Diary empty')
    else:
        st.dataframe(df.sort_values(by='timestamp', ascending=False).reset_index(drop=True))

    st.markdown('Playlist status for your user')
    user_key = username if username else 'anonymous'
    try:
        played_urls = set(df[df['username']==user_key]['youtube_url'].astype(str).tolist()) if not df.empty else set()
    except Exception:
        played_urls = set()
    st.write(f'Previously played for {user_key}: {len(played_urls)}')
    for mood, urls in YOUTUBE_PLAYLISTS.items():
        st.write(f'- {mood} ({len(urls)})')
        for u in urls:
            tag = ' (played)' if u in played_urls else ''
            st.markdown(f'   - {u}{tag}')

st.markdown('---')
st.header('Weekly Excel Summary / Export')
if st.button('Generate weekly Excel (7 days)', key='gen_excel'):
    output = generate_weekly_excel()
    st.success(f'Weekly workbook created: {output}')
    if output.exists():
        with open(output, 'rb') as f:
            st.download_button('Download weekly_summary.xlsx', f, file_name=output.name)

st.markdown('Notes: For mic recording install PyAudio (Windows: pip install pipwin; pipwin install pyaudio).')
if matched is None:
                reg_name = username.strip() if username and username.strip() else f"anon_{datetime.now().strftime('%Y%m%d%H%M%S')}"
                # Save this descriptor as a new user
                db.setdefault(reg_name, []).append(des)
                save_desc_db(db)
                st.success(f"New user registered: {reg_name}")
                user_final = reg_name
else:
                st.success(f"Recognized user: {matched} (score={score})")
                user_final = matched

            # Mood estimation
mood, conf = estimate_mood_from_face(face)
st.info(f"Mood detected: {mood} (confidence {conf})")

yt = choose_unique_youtube(user_final, mood)
wellness = pick_wellness(user_final, mood)

if yt:
                log_entry(user_final, mood, 'face', yt, wellness)
                st.video(yt)
                st.success("Playing recommended video.")
                st.info("Wellness suggestion: " + wellness)
else:
                st.warning("No YouTube URLs configured for this mood.")
