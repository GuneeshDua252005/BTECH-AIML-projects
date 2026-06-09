# # musicsoulbeatsapp_final.py
# # Streamlit app: Face/Emoji/Mic/Text -> Mood -> Auto-play YouTube (embedded)
# # Uses browser mic via streamlit-webrtc (no PyAudio required)
# # Maintains diary under entered username and exports weekly Excel with TestCases
# # Save as musicsoulbeatsapp.py and run:
# #    streamlit run musicsoulbeatsapp.py

# # Required installs (run once):
# # pip install streamlit opencv-python pandas numpy SpeechRecognition textblob openpyxl matplotlib pillow streamlit-webrtc av
# # NOTE: streamlit-webrtc pulls in aiortc/av; on Windows you may need Microsoft Build Tools or use prebuilt wheels for av.

# import streamlit as st
# import cv2
# import numpy as np
# import pandas as pd
# import os
# import pickle
# from datetime import datetime, timedelta
# from pathlib import Path
# import random
# import io
# import matplotlib.pyplot as plt
# from PIL import Image
# import base64
# import wave

# # streamlit-webrtc for browser mic capture
# from streamlit_webrtc import webrtc_streamer, WebRtcMode, AudioProcessorBase

# # optional imports
# try:
#     import speech_recognition as sr
#     SR_AVAILABLE = True
# except Exception:
#     SR_AVAILABLE = False

# try:
#     from textblob import TextBlob
#     TEXTBLOB_AVAILABLE = True
# except Exception:
#     TEXTBLOB_AVAILABLE = False

# # excel helper
# try:
#     import openpyxl
#     from openpyxl.drawing.image import Image as XLImage
#     OPENPYXL_AVAILABLE = True
# except Exception:
#     OPENPYXL_AVAILABLE = False

# st.set_page_config(page_title="SoulBeat — Final", layout="wide")

# # Data paths
# DATA_DIR = Path("soul_data")
# DATA_DIR.mkdir(exist_ok=True)
# DESC_DB = DATA_DIR / "face_descs.pkl"
# DIARY_CSV = DATA_DIR / "mood_diary.csv"
# WEEKLY_XLSX = DATA_DIR / "weekly_summary.xlsx"

# DIARY_COLUMNS = ["timestamp", "username", "mood", "source", "youtube_url", "wellness", "login_count"]

# # ORB + matcher
# ORB = cv2.ORB_create(500)
# BF = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)

# # Playlists
# YOUTUBE_PLAYLISTS = {
#     "happy": [
#         "https://www.youtube.com/watch?v=3JWTaaS7LdU",
#         "https://www.youtube.com/watch?v=d-diB65scQU",
#         "https://www.youtube.com/watch?v=2Vv-BfVoq4g"
#     ],
#     "neutral": [
#         "https://www.youtube.com/watch?v=fLexgOxsZu0",
#         "https://www.youtube.com/watch?v=5qap5aO4i9A",
#         "https://www.youtube.com/watch?v=DWcJFNfaw9c"
#     ],
#     "sad": [
#         "https://www.youtube.com/watch?v=hTWKbfoikeg",
#         "https://www.youtube.com/watch?v=RB-RcX5DS5A",
#         "https://www.youtube.com/watch?v=2XU0oxnq2qU"
#     ]
# }

# WELLNESS_SUGGESTIONS = {
#     "happy": ["Keep your energy — try a focused 15-min sprint.", "Share positivity with a friend.", "Short stretch + breathe."],
#     "neutral": ["Take a 10-min walk.", "5-min breathing.", "Hydrate and snack."],
#     "sad": ["Try 10-min guided breathing.", "Gentle stretch/yoga.", "Write one small positive thing."]
# }

# # ---------- Utilities ----------

# def ensure_diary_schema():
#     if not DIARY_CSV.exists():
#         df = pd.DataFrame(columns=DIARY_COLUMNS)
#         df.to_csv(DIARY_CSV, index=False)
#         return df
#     try:
#         df = pd.read_csv(DIARY_CSV)
#         missing = [c for c in DIARY_COLUMNS if c not in df.columns]
#         if missing:
#             for c in missing:
#                 df[c] = 0 if c == 'login_count' else ''
#             df = df.reindex(columns=DIARY_COLUMNS)
#             df.to_csv(DIARY_CSV, index=False)
#         return df
#     except Exception:
#         df = pd.DataFrame(columns=DIARY_COLUMNS)
#         df.to_csv(DIARY_CSV, index=False)
#         return df


# def load_desc_db():
#     if DESC_DB.exists():
#         try:
#             with open(DESC_DB, 'rb') as f:
#                 return pickle.load(f)
#         except Exception:
#             return {}
#     return {}


# def save_desc_db(db):
#     with open(DESC_DB, 'wb') as f:
#         pickle.dump(db, f)


# def compute_orb_descriptor(gray_face):
#     if gray_face is None or gray_face.size == 0:
#         return None
#     try:
#         h,w = gray_face.shape[:2]
#         if h < 80 or w < 80:
#             gray_face = cv2.resize(gray_face, (100,100))
#         kps, des = ORB.detectAndCompute(gray_face, None)
#         return des
#     except Exception:
#         return None


# def match_descriptor(des, db, min_matches=6):
#     if des is None:
#         return None, 0
#     best_user, best_score = None, 0
#     for user, samples in db.items():
#         for stored in samples:
#             if stored is None:
#                 continue
#             try:
#                 matches = BF.match(des, stored)
#             except Exception:
#                 continue
#             good = [m for m in matches if m.distance < 60]
#             if len(good) > best_score:
#                 best_score = len(good)
#                 best_user = user
#     if best_score >= min_matches:
#         return best_user, best_score
#     return None, best_score


# def detect_faces_haar(frame):
#     gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
#     face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
#     faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=4, minSize=(60,60))
#     boxes = []
#     for (x,y,w,h) in faces:
#         boxes.append((x,y,x+w,y+h))
#     return boxes


# def estimate_mood_from_face(face_bgr):
#     try:
#         if face_bgr is None or face_bgr.size == 0:
#             return 'neutral', 0.5
#         face = face_bgr.copy()
#         h,w = face.shape[:2]
#         if h < 80 or w < 80:
#             face = cv2.resize(face, (160,160))
#         gray = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
#         hsv = cv2.cvtColor(face, cv2.COLOR_BGR2HSV)
#         mean_v = float(np.mean(hsv[:,:,2]))
#         smile_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_smile.xml')
#         smiles = smile_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=10, minSize=(20,20))
#         if len(smiles) > 0 or mean_v > 140:
#             conf = min(1.0, mean_v/255.0 + 0.2)
#             return 'happy', round(conf,2)
#         if mean_v < 90:
#             conf = min(1.0, (90-mean_v)/90)
#             return 'sad', round(conf,2)
#         return 'neutral', round(abs(mean_v/255.0 - 0.5),2)
#     except Exception:
#         return 'neutral', 0.5


# def text_to_mood(text):
#     if not text:
#         return 'neutral'
#     if TEXTBLOB_AVAILABLE:
#         try:
#             pol = TextBlob(text).sentiment.polarity
#             if pol > 0.2:
#                 return 'happy'
#             if pol < -0.2:
#                 return 'sad'
#             return 'neutral'
#         except Exception:
#             pass
#     t = text.lower()
#     if any(w in t for w in ['sad','unhappy','depress','cry','lonely','angry','mad']):
#         return 'sad'
#     if any(w in t for w in ['happy','joy','glad','love','smile']):
#         return 'happy'
#     return 'neutral'


# def choose_unique_youtube(username, mood):
#     df = ensure_diary_schema()
#     if df.empty:
#         played = set()
#     else:
#         try:
#             played = set(df[df['username']==username]['youtube_url'].astype(str).tolist())
#         except Exception:
#             played = set()
#     candidates = YOUTUBE_PLAYLISTS.get(mood, [])
#     remaining = [u for u in candidates if u not in played]
#     if remaining:
#         return random.choice(remaining)
#     return random.choice(candidates) if candidates else None


# def pick_wellness(username, mood):
#     df = ensure_diary_schema()
#     count = 0
#     if not df.empty:
#         try:
#             count = int(df[df['username']==username].shape[0])
#         except Exception:
#             count = 0
#     pool = WELLNESS_SUGGESTIONS.get(mood, ['Take a short break and breathe.'])
#     return pool[count % len(pool)]


# def log_entry(username, mood, source, youtube_url, wellness):
#     username = username.strip() if username and username.strip() else 'anonymous'
#     df = ensure_diary_schema()
#     try:
#         login_count = int(df[df['username']==username].shape[0]) + 1 if not df.empty else 1
#     except Exception:
#         login_count = 1
#     row = {
#         'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
#         'username': username,
#         'mood': mood,
#         'source': source,
#         'youtube_url': youtube_url,
#         'wellness': wellness,
#         'login_count': login_count
#     }
#     df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
#     df.to_csv(DIARY_CSV, index=False)
#     st.session_state['last_played'] = {'username': username, 'mood': mood, 'source': source, 'youtube': youtube_url, 'time': datetime.now().isoformat()}

# # ---------- Weekly Excel ----------

# def generate_weekly_excel(output_path=WEEKLY_XLSX, days=7):
#     df = ensure_diary_schema()
#     if df.empty:
#         with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
#             pd.DataFrame(columns=DIARY_COLUMNS).to_excel(writer, sheet_name='RawLogs', index=False)
#         return output_path

#     cutoff = pd.Timestamp.now() - pd.Timedelta(days=days)
#     df['timestamp_dt'] = pd.to_datetime(df['timestamp'], errors='coerce')
#     recent = df[df['timestamp_dt'] >= cutoff].copy()
#     if recent.empty:
#         recent = pd.DataFrame(columns=df.columns)

#     if not recent.empty:
#         recent['date'] = recent['timestamp_dt'].dt.date
#         mood_counts = recent.groupby(['date','mood']).size().unstack(fill_value=0)
#         mood_daily = mood_counts.reset_index()
#         mood_totals = recent['mood'].value_counts().rename_axis('mood').reset_index(name='count')
#     else:
#         mood_daily = pd.DataFrame()
#         mood_totals = pd.DataFrame(columns=['mood','count'])

#     if not recent.empty:
#         recs = recent.groupby('wellness').size().reset_index(name='count').sort_values('count', ascending=False)
#     else:
#         recs = pd.DataFrame(columns=['wellness','count'])

#     testcases = recent.copy()
#     if not testcases.empty:
#         testcases['test_name'] = testcases['username'] + '_' + testcases['timestamp']
#         testcases['expected_mood'] = testcases['mood']
#         testcases['action'] = testcases['source']
#         testcases = testcases[['test_name','timestamp','username','action','expected_mood','youtube_url','wellness','login_count']]
#     else:
#         testcases = pd.DataFrame(columns=['test_name','timestamp','username','action','expected_mood','youtube_url','wellness','login_count'])

#     if not OPENPYXL_AVAILABLE:
#         fallback = output_path.with_suffix('.csv')
#         recent.to_csv(fallback, index=False)
#         return fallback

#     with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
#         recent.drop(columns=['timestamp_dt'], errors='ignore').to_excel(writer, sheet_name='RawLogs', index=False)
#         mood_daily.to_excel(writer, sheet_name='MoodDaily', index=False)
#         mood_totals.to_excel(writer, sheet_name='MoodTotals', index=False)
#         recs.to_excel(writer, sheet_name='Recommendations', index=False)
#         testcases.to_excel(writer, sheet_name='TestCases', index=False)

#     try:
#         fig1, ax1 = plt.subplots(figsize=(6,4))
#         if not mood_totals.empty:
#             ax1.bar(mood_totals['mood'], mood_totals['count'])
#             ax1.set_title('Mood distribution (last {} days)'.format(days))
#             ax1.set_ylabel('Count')
#         else:
#             ax1.text(0.5,0.5,'No data',ha='center')
#         img1 = output_path.parent / 'chart1.png'
#         fig1.savefig(img1)
#         plt.close(fig1)

#         fig2, ax2 = plt.subplots(figsize=(8,4))
#         if not mood_daily.empty:
#             mood_daily.set_index('date').plot(kind='bar', stacked=True, ax=ax2)
#             ax2.set_title('Mood counts by day')
#         else:
#             ax2.text(0.5,0.5,'No data',ha='center')
#         img2 = output_path.parent / 'chart2.png'
#         fig2.savefig(img2)
#         plt.close(fig2)

#         wb = openpyxl.load_workbook(output_path)
#         ws = wb.create_sheet('Charts')
#         img_obj1 = XLImage(str(img1))
#         img_obj2 = XLImage(str(img2))
#         ws.add_image(img_obj1, 'A1')
#         ws.add_image(img_obj2, 'A25')
#         wb.save(output_path)

#         try:
#             img1.unlink()
#             img2.unlink()
#         except Exception:
#             pass
#     except Exception:
#         pass

#     return output_path

# # ---------- Playback helper ----------

# def safe_play_and_log(username, mood, source):
#     yt = choose_unique_youtube(username, mood)
#     wellness = pick_wellness(username, mood)
#     last = st.session_state.get('last_played')
#     now = datetime.now()
#     if last:
#         try:
#             dt = datetime.fromisoformat(last.get('time'))
#             same = (last.get('username') == username and last.get('mood') == mood and last.get('source') == source)
#             if same and (now - dt).total_seconds() < 3:
#                 if yt:
#                     st.video(yt)
#                     st.info('Wellness: ' + wellness)
#                 return
#         except Exception:
#             pass
#     if yt:
#         st.video(yt)
#         st.info('Wellness: ' + wellness)
#         log_entry(username, mood, source, yt, wellness)
#     else:
#         st.warning('No yt configured for this mood')

# # ---------- Browser mic: streamlit-webrtc audio processor ----------

# class Recorder(AudioProcessorBase):
#     """Collects microphone audio chunks from browser and stores as WAV in session_state."""
#     def __init__(self):
#         self.frames = []

#     def recv(self, frame):
#         # frame is av.AudioFrame
#         # convert to numpy array
#         try:
#             arr = frame.to_ndarray()
#             # arr shape: (n_channels, n_samples) or (n_samples,) depending
#             self.frames.append(arr)
#         except Exception:
#             pass
#         return frame


# def assemble_wav_from_frames(frames, sample_rate=48000):
#     # frames: list of numpy arrays (n_channels, n_samples)
#     # convert to mono 16-bit PCM and return bytes
#     if not frames:
#         return None
#     # concat along samples
#     try:
#         # normalize shapes
#         parts = []
#         for a in frames:
#             if a.ndim == 2:
#                 # take mean across channels
#                 a_mono = a.mean(axis=0).astype(np.int16)
#             else:
#                 a_mono = a.astype(np.int16)
#             parts.append(a_mono)
#         data = np.concatenate(parts)
#         # write to WAV bytes
#         buf = io.BytesIO()
#         with wave.open(buf, 'wb') as wf:
#             wf.setnchannels(1)
#             wf.setsampwidth(2)  # 16-bit
#             wf.setframerate(sample_rate)
#             wf.writeframes(data.tobytes())
#         return buf.getvalue()
#     except Exception:
#         return None


# def transcribe_audio_bytes(audio_bytes):
#     if not SR_AVAILABLE:
#         return None, 'SpeechRecognition not installed'
#     try:
#         r = sr.Recognizer()
#         with sr.AudioFile(io.BytesIO(audio_bytes)) as source:
#             audio = r.record(source)
#         # use Google recognizer (no API key) - supports language param if needed
#         text = r.recognize_google(audio)
#         return text, None
#     except sr.UnknownValueError:
#         return None, 'Could not understand audio'
#     except Exception as e:
#         return None, str(e)

# # ---------- Streamlit UI ----------
# st.title('SoulBeat — Final (Browser Mic, Auto-play, Diary, Excel)')

# st.sidebar.header('User & Settings')
# username = st.sidebar.text_input('Username (for diary):', value=st.session_state.get('username',''), key='username')
# if not username or not username.strip():
#     st.sidebar.warning('Please enter your name — diary will use this username (no anonymous).')

# lang = st.sidebar.selectbox('Speech language', ['English','Hindi'], index=0)
# lang_code = 'en-US' if lang=='English' else 'hi-IN'

# st.sidebar.markdown('Run with: streamlit run musicsoulbeatsapp.py')

# col1, col2 = st.columns([1,2])

# with col1:
#     st.header('Manual Controls')
#     emoji_map = {'😞':'sad','😐':'neutral','🙂':'happy','😊':'happy','😡':'sad'}
#     emoji_choice = st.select_slider('Pick an emoji', options=list(emoji_map.keys()), value='😐', key='emoji_slider')
#     manual_mood = emoji_map.get(emoji_choice,'neutral')
#     # auto-play when emoji changes
#     if st.session_state.get('emoji_last') != emoji_choice and username and username.strip():
#         st.session_state['emoji_last'] = emoji_choice
#         safe_play_and_log(username.strip(), manual_mood, 'emoji')

#     typed = st.text_input('Or type mood (happy/neutral/sad):','', key='typed_mood')
#     if typed and typed.strip().lower() in {'happy','neutral','sad'} and username and username.strip():
#         safe_play_and_log(username.strip(), typed.strip().lower(), 'typed')

#     st.markdown('---')
#     st.header('Speech Input (Browser Microphone)')
#     st.write('Click Start to stream your browser microphone. Speak; then click "Capture & Transcribe" to process a recent short buffer.')

#     webrtc_ctx = webrtc_streamer(key='webrtc', mode=WebRtcMode.SENDONLY, audio_processor_factory=Recorder,
#                                  media_stream_constraints={"audio": True, "video": False}, async_processing=True)

#     if st.button('Capture & Transcribe (browser mic)', key='capture_transcribe'):
#         if not username or not username.strip():
#             st.error('Enter username first in sidebar')
#         else:
#             # attempt to fetch frames from the processor
#             processor = webrtc_ctx.audio_processor
#             if processor is None:
#                 st.error('No audio processor available. Start the microphone stream first.')
#             else:
#                 frames = []
#                 try:
#                     # grab frames collected so far
#                     frames = processor.frames[-80:]  # last N chunks
#                 except Exception:
#                     frames = []
#                 audio_bytes = assemble_wav_from_frames(frames, sample_rate=48000)
#                 if not audio_bytes:
#                     st.error('No audio data captured. Make sure microphone access is allowed and speak for a few seconds before capturing.')
#                 else:
#                     # Write temp wav for user visibility
#                     st.audio(audio_bytes)
#                     # transcribe
#                     if SR_AVAILABLE:
#                         try:
#                             r = sr.Recognizer()
#                             with sr.AudioFile(io.BytesIO(audio_bytes)) as source:
#                                 audio = r.record(source)
#                             text = r.recognize_google(audio, language=lang_code)
#                             st.write('Transcribed:', text)
#                             mood = text_to_mood(text)
#                             st.write('Detected mood:', mood)
#                             safe_play_and_log(username.strip(), mood, 'browser_mic')
#                         except Exception as e:
#                             st.error('Transcription failed: ' + str(e))
#                     else:
#                         st.info('SpeechRecognition not installed — cannot transcribe. Install SpeechRecognition for transcription.')

#     st.markdown('---')
#     st.header('Upload audio')
#     uploaded = st.file_uploader('Upload audio (wav/mp3/m4a)', type=['wav','mp3','m4a'], key='upload_audio')
#     if uploaded is not None and st.button('Transcribe upload and play', key='transcribe_upload'):
#         if not username or not username.strip():
#             st.error('Enter username first in sidebar')
#         else:
#             tmp = DATA_DIR / f"tmp_{datetime.now().strftime('%Y%m%d%H%M%S')}"
#             ext = uploaded.name.split('.')[-1]
#             tmp = tmp.with_suffix('.' + ext)
#             with open(tmp, 'wb') as f:
#                 f.write(uploaded.read())
#             if SR_AVAILABLE:
#                 try:
#                     r = sr.Recognizer()
#                     with sr.AudioFile(str(tmp)) as source:
#                         audio = r.record(source)
#                     text = r.recognize_google(audio, language=lang_code)
#                     st.write('Transcribed:', text)
#                     mood = text_to_mood(text)
#                     st.write('Detected mood:', mood)
#                     safe_play_and_log(username.strip(), mood, 'upload_audio')
#                 except Exception as e:
#                     st.error('Transcription error: ' + str(e))
#                 finally:
#                     try:
#                         tmp.unlink()
#                     except Exception:
#                         pass
#             else:
#                 st.info('Install SpeechRecognition to enable transcription.')

#     st.markdown('---')
#     st.header('Manual Playback')
#     if st.button('Play for manual mood', key='manual_play'):
#         if not username or not username.strip():
#             st.error('Enter username first in sidebar')
#         else:
#             final_mood = (st.session_state.get('typed_mood') or manual_mood).strip().lower()
#             if final_mood not in {'happy','neutral','sad'}:
#                 st.error('Manual mood invalid')
#             else:
#                 safe_play_and_log(username.strip(), final_mood, 'manual')

# with col2:
#     st.header('Face capture & recognition (login)')
#     st.write('Capture a photo — app detects face, registers/recognizes user using the entered username, estimates mood, auto-plays and logs the entry.')
#     img = st.camera_input('Capture photo', key='camera')
#     if img is not None:
#         if not username or not username.strip():
#             st.error('Enter username first in sidebar')
#         else:
#             arr = np.frombuffer(img.getvalue(), np.uint8)
#             frame = cv2.imdecode(arr, cv2.IMREAD_COLOR)
#             st.image(frame, channels='BGR')
#             boxes = detect_faces_haar(frame)
#             if not boxes:
#                 st.warning('No face detected')
#             else:
#                 x1,y1,x2,y2 = boxes[0]
#                 x1, y1 = max(0,x1), max(0,y1)
#                 x2, y2 = min(frame.shape[1], x2), min(frame.shape[0], y2)
#                 face = frame[y1:y2, x1:x2]
#                 st.image(face, channels='BGR', caption='Face crop')
#                 gray = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
#                 des = compute_orb_descriptor(gray)
#                 db = load_desc_db()
#                 matched, score = match_descriptor(des, db, min_matches=6)
#                 if matched is None:
#                     reg_name = username.strip()
#                     db.setdefault(reg_name, []).append(des)
#                     save_desc_db(db)
#                     st.info(f'Registered descriptor for {reg_name}')
#                     userkey = reg_name
#                 else:
#                     userkey = matched
#                     db.setdefault(userkey, []).append(des)
#                     save_desc_db(db)
#                     st.success(f'Recognized {userkey} (score {score})')
#                 mood, conf = estimate_mood_from_face(face)
#                 st.write(f'Detected mood: {mood} (conf {conf})')
#                 safe_play_and_log(userkey, mood, 'face_capture')

#     st.markdown('---')
#     st.header('Diary & Playlists')
#     df = ensure_diary_schema()
#     if df.empty:
#         st.info('Diary empty')
#     else:
#         st.dataframe(df.sort_values(by='timestamp', ascending=False).reset_index(drop=True))

#     st.markdown('Playlist status for your user')
#     user_key = username.strip() if username and username.strip() else 'anonymous'
#     try:
#         played_urls = set(df[df['username']==user_key]['youtube_url'].astype(str).tolist()) if not df.empty else set()
#     except Exception:
#         played_urls = set()
#     st.write(f'Previously played for {user_key}: {len(played_urls)}')
#     for mood, urls in YOUTUBE_PLAYLISTS.items():
#         st.write(f'- {mood} ({len(urls)})')
#         for u in urls:
#             tag = ' (played)' if u in played_urls else ''
#             st.markdown(f'   - {u}{tag}')

# st.markdown('---')
# st.header('Weekly Excel Summary / Export')
# if st.button('Generate weekly Excel (7 days)', key='gen_excel'):
#     out = generate_weekly_excel()
#     st.success(f'Weekly workbook created: {out}')
#     if out.exists():
#         with open(out, 'rb') as f:
#             st.download_button('Download weekly_summary.xlsx', f, file_name=out.name)

# st.markdown('Notes: Use the browser mic control above to stream your microphone (no PyAudio required). Install SpeechRecognition to enable transcription for uploaded or captured audio.')
# musicsoulbeatsapp.py
# SoulBeat: face/emoji/mic/text -> mood -> auto-play YouTube + diary + weekly Excel (TestCases)
# Save & run: streamlit run musicsoulbeatsapp.py

import streamlit as st
import cv2
import numpy as np
import pandas as pd
import pickle
from datetime import datetime, timedelta
from pathlib import Path
import random
import io
import matplotlib.pyplot as plt
from PIL import Image
import wave

# Browser mic via streamlit-webrtc
from streamlit_webrtc import webrtc_streamer, WebRtcMode, AudioProcessorBase

# optional imports
try:
    import speech_recognition as sr
    SR_AVAILABLE = True
except Exception:
    SR_AVAILABLE = False

try:
    from textblob import TextBlob
    TEXTBLOB_AVAILABLE = True
except Exception:
    TEXTBLOB_AVAILABLE = False

# excel helper
try:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

st.set_page_config(page_title="SoulBeat — Final", layout="wide")

# Data paths
DATA_DIR = Path("soul_data")
DATA_DIR.mkdir(exist_ok=True)
DESC_DB = DATA_DIR / "face_descs.pkl"
DIARY_CSV = DATA_DIR / "mood_diary.csv"
WEEKLY_XLSX = DATA_DIR / "weekly_summary.xlsx"

DIARY_COLUMNS = ["timestamp", "username", "mood", "source", "youtube_url", "wellness", "login_count"]

# ORB + matcher
ORB = cv2.ORB_create(500)
BF = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)

# Playlists (added angry)
YOUTUBE_PLAYLISTS = {
    "happy": [
        "https://www.youtube.com/watch?v=3JWTaaS7LdU",
        "https://www.youtube.com/watch?v=d-diB65scQU",
        "https://www.youtube.com/watch?v=2Vv-BfVoq4g"
    ],
    "neutral": [
        "https://www.youtube.com/watch?v=fLexgOxsZu0",
        "https://www.youtube.com/watch?v=5qap5aO4i9A",
        "https://www.youtube.com/watch?v=DWcJFNfaw9c"
    ],
    "sad": [
        "https://www.youtube.com/watch?v=hTWKbfoikeg",
        "https://www.youtube.com/watch?v=RB-RcX5DS5A",
        "https://www.youtube.com/watch?v=2XU0oxnq2qU"
    ],
    "angry": [   # new angry playlist
        "https://www.youtube.com/watch?v=2vjPBrBU-TM",
        "https://www.youtube.com/watch?v=6Ejga4kJUts",
        "https://www.youtube.com/watch?v=Zi_XLOBDo_Y"
    ]
}

WELLNESS_SUGGESTIONS = {
    "happy": ["Keep your energy — try a focused 15-min sprint.", "Share positivity with a friend.", "Short stretch + breathe."],
    "neutral": ["Take a 10-min walk.", "5-min breathing.", "Hydrate and snack."],
    "sad": ["Try 10-min guided breathing.", "Gentle stretch/yoga.", "Write one small positive thing."],
    "angry": ["Pause and breathe deeply for 1-2 minutes.", "Take a short walk to calm down.", "Count to ten and drink water."]
}

# ---------- Utilities ----------
def ensure_diary_schema():
    if not DIARY_CSV.exists():
        df = pd.DataFrame(columns=DIARY_COLUMNS)
        df.to_csv(DIARY_CSV, index=False)
        return df
    try:
        df = pd.read_csv(DIARY_CSV)
        missing = [c for c in DIARY_COLUMNS if c not in df.columns]
        if missing:
            for c in missing:
                df[c] = 0 if c == "login_count" else ""
            df = df.reindex(columns=DIARY_COLUMNS)
            df.to_csv(DIARY_CSV, index=False)
        return df
    except Exception:
        df = pd.DataFrame(columns=DIARY_COLUMNS)
        df.to_csv(DIARY_CSV, index=False)
        return df

def load_desc_db():
    if DESC_DB.exists():
        try:
            with open(DESC_DB, 'rb') as f:
                return pickle.load(f)
        except Exception:
            return {}
    return {}

def save_desc_db(db):
    with open(DESC_DB, 'wb') as f:
        pickle.dump(db, f)

def compute_orb_descriptor(gray_face):
    if gray_face is None or gray_face.size == 0:
        return None
    try:
        h,w = gray_face.shape[:2]
        if h < 80 or w < 80:
            gray_face = cv2.resize(gray_face, (100,100))
        kps, des = ORB.detectAndCompute(gray_face, None)
        return des
    except Exception:
        return None

def match_descriptor(des, db, min_matches=6):
    if des is None:
        return None, 0
    best_user, best_score = None, 0
    for user, samples in db.items():
        for stored in samples:
            if stored is None:
                continue
            try:
                matches = BF.match(des, stored)
            except Exception:
                continue
            good = [m for m in matches if m.distance < 60]
            if len(good) > best_score:
                best_score = len(good)
                best_user = user
    if best_score >= min_matches:
        return best_user, best_score
    return None, best_score

def detect_faces_haar(frame):
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=4, minSize=(60,60))
    boxes = []
    for (x,y,w,h) in faces:
        boxes.append((x,y,x+w,y+h))
    return boxes

def estimate_mood_from_face(face_bgr):
    try:
        if face_bgr is None or face_bgr.size == 0:
            return 'neutral', 0.5
        face = face_bgr.copy()
        h,w = face.shape[:2]
        if h < 80 or w < 80:
            face = cv2.resize(face, (160,160))
        gray = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
        hsv = cv2.cvtColor(face, cv2.COLOR_BGR2HSV)
        mean_v = float(np.mean(hsv[:,:,2]))
        smile_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_smile.xml')
        smiles = smile_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=10, minSize=(20,20))
        if len(smiles) > 0 or mean_v > 140:
            conf = min(1.0, mean_v/255.0 + 0.2)
            return 'happy', round(conf,2)
        if mean_v < 90:
            conf = min(1.0, (90-mean_v)/90)
            return 'sad', round(conf,2)
        return 'neutral', round(abs(mean_v/255.0 - 0.5),2)
    except Exception:
        return 'neutral', 0.5

def text_to_mood(text):
    if not text:
        return 'neutral'
    if TEXTBLOB_AVAILABLE:
        try:
            pol = TextBlob(text).sentiment.polarity
            if pol > 0.2:
                return 'happy'
            if pol < -0.2:
                return 'sad'
            return 'neutral'
        except Exception:
            pass
    t = text.lower()
    if any(w in t for w in ['sad','unhappy','depress','cry','lonely','angry','mad','furious']):
        return 'sad' if 'sad' in t else ('angry' if any(a in t for a in ['angry','mad','furious']) else 'sad')
    if any(w in t for w in ['happy','joy','glad','love','smile']):
        return 'happy'
    return 'neutral'

def choose_unique_youtube(username, mood):
    df = ensure_diary_schema()
    if df.empty:
        played = set()
    else:
        try:
            played = set(df[df['username']==username]['youtube_url'].astype(str).tolist())
        except Exception:
            played = set()
    candidates = YOUTUBE_PLAYLISTS.get(mood, [])
    remaining = [u for u in candidates if u not in played]
    if remaining:
        return random.choice(remaining)
    return random.choice(candidates) if candidates else None

def pick_wellness(username, mood):
    df = ensure_diary_schema()
    count = 0
    if not df.empty:
        try:
            count = int(df[df['username']==username].shape[0])
        except Exception:
            count = 0
    pool = WELLNESS_SUGGESTIONS.get(mood, ['Take a short break and breathe.'])
    return pool[count % len(pool)]

def log_entry(username, mood, source, youtube_url, wellness):
    username = username.strip()
    df = ensure_diary_schema()
    try:
        login_count = int(df[df['username']==username].shape[0]) + 1 if not df.empty else 1
    except Exception:
        login_count = 1
    row = {
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'username': username,
        'mood': mood,
        'source': source,
        'youtube_url': youtube_url,
        'wellness': wellness,
        'login_count': login_count
    }
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    df.to_csv(DIARY_CSV, index=False)
    # store last played to avoid duplicates
    st.session_state['last_played'] = {'username': username, 'mood': mood, 'source': source, 'youtube': youtube_url, 'time': datetime.now().isoformat()}

# ---------- Weekly Excel ----------
def generate_weekly_excel(output_path=WEEKLY_XLSX, days=7):
    df = ensure_diary_schema()
    if df.empty:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            pd.DataFrame(columns=DIARY_COLUMNS).to_excel(writer, sheet_name='RawLogs', index=False)
        return output_path

    cutoff = pd.Timestamp.now() - pd.Timedelta(days=days)
    df['timestamp_dt'] = pd.to_datetime(df['timestamp'], errors='coerce')
    recent = df[df['timestamp_dt'] >= cutoff].copy()
    if recent.empty:
        recent = pd.DataFrame(columns=df.columns)

    if not recent.empty:
        recent['date'] = recent['timestamp_dt'].dt.date
        mood_counts = recent.groupby(['date','mood']).size().unstack(fill_value=0)
        mood_daily = mood_counts.reset_index()
        mood_totals = recent['mood'].value_counts().rename_axis('mood').reset_index(name='count')
    else:
        mood_daily = pd.DataFrame()
        mood_totals = pd.DataFrame(columns=['mood','count'])

    if not recent.empty:
        recs = recent.groupby('wellness').size().reset_index(name='count').sort_values('count', ascending=False)
    else:
        recs = pd.DataFrame(columns=['wellness','count'])

    # TestCases sheet generation
    testcases = recent.copy()
    if not testcases.empty:
        testcases['test_name'] = testcases['username'] + '_' + testcases['timestamp']
        testcases['expected_mood'] = testcases['mood']
        testcases['action'] = testcases['source']
        testcases = testcases[['test_name','timestamp','username','action','expected_mood','youtube_url','wellness','login_count']]
    else:
        testcases = pd.DataFrame(columns=['test_name','timestamp','username','action','expected_mood','youtube_url','wellness','login_count'])

    if not OPENPYXL_AVAILABLE:
        fallback = output_path.with_suffix('.csv')
        recent.to_csv(fallback, index=False)
        return fallback

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        recent.drop(columns=['timestamp_dt'], errors='ignore').to_excel(writer, sheet_name='RawLogs', index=False)
        mood_daily.to_excel(writer, sheet_name='MoodDaily', index=False)
        mood_totals.to_excel(writer, sheet_name='MoodTotals', index=False)
        recs.to_excel(writer, sheet_name='Recommendations', index=False)
        testcases.to_excel(writer, sheet_name='TestCases', index=False)

    # embed charts
    try:
        fig1, ax1 = plt.subplots(figsize=(6,4))
        if not mood_totals.empty:
            ax1.bar(mood_totals['mood'], mood_totals['count'])
            ax1.set_title('Mood distribution (last {} days)'.format(days))
            ax1.set_ylabel('Count')
        else:
            ax1.text(0.5,0.5,'No data',ha='center')
        img1 = output_path.parent / 'chart1.png'
        fig1.savefig(img1)
        plt.close(fig1)

        fig2, ax2 = plt.subplots(figsize=(8,4))
        if not mood_daily.empty:
            mood_daily.set_index('date').plot(kind='bar', stacked=True, ax=ax2)
            ax2.set_title('Mood counts by day')
        else:
            ax2.text(0.5,0.5,'No data',ha='center')
        img2 = output_path.parent / 'chart2.png'
        fig2.savefig(img2)
        plt.close(fig2)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.create_sheet('Charts')
        img_obj1 = XLImage(str(img1))
        img_obj2 = XLImage(str(img2))
        ws.add_image(img_obj1, 'A1')
        ws.add_image(img_obj2, 'A25')
        wb.save(output_path)

        try:
            img1.unlink()
            img2.unlink()
        except Exception:
            pass
    except Exception:
        pass

    return output_path

# ---------- Playback helper ----------
def safe_play_and_log(username, mood, source):
    username = username.strip()
    yt = choose_unique_youtube(username, mood)
    wellness = pick_wellness(username, mood)
    last = st.session_state.get('last_played')
    now = datetime.now()
    if last:
        try:
            dt = datetime.fromisoformat(last.get('time'))
            same = (last.get('username') == username and last.get('mood') == mood and last.get('source') == source)
            if same and (now - dt).total_seconds() < 3:
                if yt:
                    st.video(yt)
                    st.info('Wellness: ' + wellness)
                return
        except Exception:
            pass
    if yt:
        st.video(yt)
        st.info('Wellness: ' + wellness)
        log_entry(username, mood, source, yt, wellness)
    else:
        st.warning('No yt configured for this mood')

# ---------- Browser mic: streamlit-webrtc audio processor ----------
class Recorder(AudioProcessorBase):
    def __init__(self):
        self.frames = []
    def recv(self, frame):
        try:
            arr = frame.to_ndarray()
            self.frames.append(arr)
        except Exception:
            pass
        return frame

def assemble_wav_from_frames(frames, sample_rate=48000):
    if not frames:
        return None
    try:
        parts = []
        for a in frames:
            if a.ndim == 2:
                a_mono = a.mean(axis=0).astype(np.int16)
            else:
                a_mono = a.astype(np.int16)
            parts.append(a_mono)
        data = np.concatenate(parts)
        buf = io.BytesIO()
        with wave.open(buf, 'wb') as wf:
            wf.setnchannels(1)
            wf.setsampwidth(2)
            wf.setframerate(sample_rate)
            wf.writeframes(data.tobytes())
        return buf.getvalue()
    except Exception:
        return None

# ---------- UI ----------
st.title('SoulBeat — Final (Browser Mic, Auto-play, Diary, Excel)')

st.sidebar.header('User & Settings')
username = st.sidebar.text_input('Username (for diary):', value=st.session_state.get('username',''), key='username')
if not username or not username.strip():
    st.sidebar.warning('Enter your name — diary will record entries under this username.')

lang = st.sidebar.selectbox('Speech language', ['English','Hindi'], index=0)
lang_code = 'en-US' if lang == 'English' else 'hi-IN'

st.sidebar.markdown('Run with: streamlit run musicsoulbeatsapp.py')

col1, col2 = st.columns([1,2])

with col1:
    st.header('Manual Controls')
    emoji_map = {'😞':'sad','😐':'neutral','🙂':'happy','😊':'happy','😡':'angry'}
    emoji_choice = st.select_slider('Pick an emoji', options=list(emoji_map.keys()), value='😐', key='emoji_slider')
    manual_mood = emoji_map.get(emoji_choice,'neutral')
    # auto-play when emoji changes (requires username)
    if st.session_state.get('emoji_last') != emoji_choice and username and username.strip():
        st.session_state['emoji_last'] = emoji_choice
        safe_play_and_log(username.strip(), manual_mood, 'emoji')

    typed = st.text_input('Or type mood (happy/neutral/sad/angry):','', key='typed_mood')
    if typed and typed.strip().lower() in {'happy','neutral','sad','angry'} and username and username.strip():
        safe_play_and_log(username.strip(), typed.strip().lower(), 'typed')

    st.markdown('---')
    st.header('Speech Input (Browser Microphone)')
    st.write('Start the microphone stream, speak, then click "Capture & Transcribe" to process recent audio.')

    webrtc_ctx = webrtc_streamer(key='webrtc', mode=WebRtcMode.SENDONLY, audio_processor_factory=Recorder,
                                 media_stream_constraints={"audio": True, "video": False}, async_processing=True)

    if st.button('Capture & Transcribe (browser mic)', key='capture_transcribe'):
        if not username or not username.strip():
            st.error('Enter username first in sidebar')
        else:
            processor = webrtc_ctx.audio_processor
            if processor is None:
                st.error('Start microphone stream first')
            else:
                frames = []
                try:
                    frames = processor.frames[-120:]
                except Exception:
                    frames = []
                audio_bytes = assemble_wav_from_frames(frames, sample_rate=48000)
                if not audio_bytes:
                    st.error('No audio captured — speak into your mic then capture.')
                else:
                    st.audio(audio_bytes)
                    if SR_AVAILABLE:
                        try:
                            r = sr.Recognizer()
                            with sr.AudioFile(io.BytesIO(audio_bytes)) as source:
                                audio = r.record(source)
                            text = r.recognize_google(audio, language=lang_code)
                            st.write('Transcribed:', text)
                            mood = text_to_mood(text)
                            st.write('Detected mood:', mood)
                            safe_play_and_log(username.strip(), mood, 'browser_mic')
                        except Exception as e:
                            st.error('Transcription failed: ' + str(e))
                    else:
                        st.info('SpeechRecognition not installed — cannot transcribe.')

    st.markdown('---')
    st.header('Upload audio')
    uploaded = st.file_uploader('Upload audio (wav/mp3/m4a)', type=['wav','mp3','m4a'], key='upload_audio')
    if uploaded is not None and st.button('Transcribe upload and play', key='transcribe_upload'):
        if not username or not username.strip():
            st.error('Enter username first')
        else:
            tmp = DATA_DIR / f"tmp_{datetime.now().strftime('%Y%m%d%H%M%S')}"
            ext = uploaded.name.split('.')[-1]
            tmp = tmp.with_suffix('.' + ext)
            with open(tmp, 'wb') as f:
                f.write(uploaded.read())
            if SR_AVAILABLE:
                try:
                    r = sr.Recognizer()
                    with sr.AudioFile(str(tmp)) as source:
                        audio = r.record(source)
                    text = r.recognize_google(audio, language=lang_code)
                    st.write('Transcribed:', text)
                    mood = text_to_mood(text)
                    st.write('Detected mood:', mood)
                    safe_play_and_log(username.strip(), mood, 'upload_audio')
                except Exception as e:
                    st.error('Transcription error: ' + str(e))
                finally:
                    try:
                        tmp.unlink()
                    except Exception:
                        pass
            else:
                st.info('Install SpeechRecognition for transcription.')

    st.markdown('---')
    st.header('Manual Playback')
    if st.button('Play for manual mood', key='manual_play'):
        if not username or not username.strip():
            st.error('Enter username first')
        else:
            final_mood = (st.session_state.get('typed_mood') or manual_mood).strip().lower()
            if final_mood not in {'happy','neutral','sad','angry'}:
                st.error('Manual mood invalid')
            else:
                safe_play_and_log(username.strip(), final_mood, 'manual')

with col2:
    st.header('Face capture & recognition (login)')
    st.write('Capture photo — app will detect face, register/recognize user using entered username, estimate mood, auto-play and log entry.')
    img = st.camera_input('Capture photo', key='camera')
    if img is not None:
        if not username or not username.strip():
            st.error('Enter username first in sidebar')
        else:
            arr = np.frombuffer(img.getvalue(), np.uint8)
            frame = cv2.imdecode(arr, cv2.IMREAD_COLOR)
            st.image(frame, channels='BGR')
            boxes = detect_faces_haar(frame)
            if not boxes:
                st.warning('No face detected')
            else:
                x1,y1,x2,y2 = boxes[0]
                x1, y1 = max(0,x1), max(0,y1)
                x2, y2 = min(frame.shape[1], x2), min(frame.shape[0], y2)
                face = frame[y1:y2, x1:x2]
                st.image(face, channels='BGR', caption='Face crop')
                gray = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
                des = compute_orb_descriptor(gray)
                db = load_desc_db()
                matched, score = match_descriptor(des, db, min_matches=6)
                if matched is None:
                    reg_name = username.strip()
                    db.setdefault(reg_name, []).append(des)
                    save_desc_db(db)
                    st.info(f'Registered descriptor for {reg_name}')
                    userkey = reg_name
                else:
                    userkey = matched
                    db.setdefault(userkey, []).append(des)
                    save_desc_db(db)
                    st.success(f'Recognized {userkey} (score {score})')
                mood, conf = estimate_mood_from_face(face)
                st.write(f'Detected mood: {mood} (conf {conf})')
                safe_play_and_log(userkey, mood, 'face_capture')

    st.markdown('---')
    st.header('Diary & Playlists')
    df = ensure_diary_schema()
    if df.empty:
        st.info('Diary empty')
    else:
        st.dataframe(df.sort_values(by='timestamp', ascending=False).reset_index(drop=True))

    st.markdown('Playlist status for your user')
    user_key = username.strip() if username and username.strip() else ''
    try:
        played_urls = set(df[df['username']==user_key]['youtube_url'].astype(str).tolist()) if not df.empty and user_key else set()
    except Exception:
        played_urls = set()
    st.write(f'- {mood} ({len(urls)})')
    for u in urls:
            tag = ' (played)' if u in played_urls else ''
            st.markdown(f'   - {u}{tag}')

st.markdown('---')
st.header('Weekly Excel Summary / Export')
if st.button('Generate weekly Excel (7 days)', key='gen_excel'):
    out = generate_weekly_excel()
    st.success(f'Weekly workbook created: {out}')
    if out.exists():
        with open(out, 'rb') as f:
            st.download_button('Download weekly_summary.xlsx', f, file_name=out.name)

st.markdown('Notes: Browser mic used (no PyAudio). Install SpeechRecognition to enable transcription for uploads/capture.')
